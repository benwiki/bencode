import os
from openpyxl import Workbook  # type: ignore
from openpyxl.drawing.image import Image as ExcelImage  # type: ignore
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
from PIL import Image

def create_image_excel(folder_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    if ws is None:
        print("Error: Could not create a new worksheet.")
        return

    # Set the column headers
    ws['A1'] = 'id'
    ws['B1'] = 'image'

    # Initialize row counter
    row_num = 2

    # Iterate over all files in the folder
    for filename in os.listdir(folder_path):
        if filename.lower().endswith('.jpg'):
            # Extract the image name without extension
            image_id = os.path.splitext(filename)[0]

            # Load the image using PIL
            img_path = os.path.join(folder_path, filename)
            pil_img = Image.open(img_path)

            # Convert PIL image to Excel image
            excel_img = ExcelImage(pil_img)

            # Add the image ID to the first column
            ws.cell(row=row_num, column=1, value=image_id)

            # Add the image to the second column
            ws.add_image(excel_img, f'B{row_num}')

            # Increment the row counter
            row_num += 1

    # Define the table range (from A1 to B{last_row})
    table_range = f"A1:B{row_num - 1}"

    # Create a table
    table = Table(displayName="ImageTable", ref=table_range)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Save the workbook to the specified path
    excel_path = os.path.join(folder_path, 'images.xlsx')
    wb.save(excel_path)

    print(f"Excel file saved at {excel_path}")

# Example usage:
create_image_excel(r"C:\Users\b.hargitai\Downloads\pl_extern_how_to_2025-01-24")